import subprocess
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook

from descriptionate import parse_srt, generate_description, generate_title
from schedule_utils import next_slots 

# ──────────────────────────── Config ─────────────────────────────
output_dir = "outputvid"
excel_path = Path("videouploader.xlsx")
subtitle_dir = Path("C:\\Users\\thoma\\OneDrive\\Documents\\code_projects\\git_proj\\Bulk_Uploader\\src\\subtitles") ######################AANPASSEN################
youtube_tags = "#shorts #youtubeshorts #viral #trending"
tiktok_tags  = "#fy #foryou #tiktok #viral #fyp"
insta_tags   = "#fy #foryou #reels #viral #instareels"

# ────────────────────────── Validatie ────────────────────────────
if len(sys.argv) < 3:
    sys.exit("Usage: python main.py '<YOUTUBE_URL>' <GEMINI_API_KEY>")

url, api_key = sys.argv[1:3]

# ───────────────────── Video‑extractie & combiner ────────────────
subprocess.run(["node", "main_combiner.js", url], check=True)

# ───────────────────────── Excel voorbereiden ────────────────────
if excel_path.exists():
    wb = load_workbook(excel_path)
    ws = wb.active
    last_dt = max(
        datetime.strptime(r[0].value, "%Y-%m-%d %H:%M")
        for r in ws.iter_rows(min_row=2, max_col=1)
        if r[0].value
    )
    start_dt = last_dt + timedelta(minutes=1)
else:
    wb = Workbook()
    ws = wb.active
    ws.append(  # Publer‑kolommen
        [
            "Date",
            "Text",
            "Link",
            "Media URL(s)",
            "Title",
            "Label(s)",
            "Alt text(s)",
            "Comment(s)",
            "Pin board / FB album / Google category",
            "Post Subtype",
            "CTA",
            "Reminder",
        ]
    )
    start_dt = datetime.now()

slot_iter = next_slots(start_dt)

# ───────────────────────── Main‑loop ─────────────────────────────
for filename in os.listdir(output_dir):
    file_path = Path(output_dir) / filename
    if not file_path.is_file():
        continue

    # Genereer .srt (caption_gen.py) & lees het
    subprocess.run(
        ["python", "caption_gen.py", str(file_path), file_path.stem], check=True
    )
    srt_path = subtitle_dir / f"{file_path.stem}.srt"
    subs = parse_srt(srt_path)

    # Titel & beschrijving
    title = generate_title(subs, api_key)
    descr = generate_description(subs, api_key)

    # Plak ondertiteling op de video
    proc_captioning = subprocess.run(['python', 'subtitler.py',
                                          str(file_path),
                                          srt_path,
                                         '-o', f'final\\{file_path.stem}_subtitled.mp4',
                                         '--position', 'middle'],
                                         capture_output=True, text=True)


    # Platform & planning
    platform, when_dt = next(slot_iter)
    extra_tags = {
        "YouTube Shorts": youtube_tags,
        "TikTok": tiktok_tags,
        "Instagram Reels": insta_tags,
    }[platform]
    full_descr = f"{descr}\n\n{extra_tags}"

    # Rij toevoegen
    ws.append(
        [
            when_dt.strftime("%Y-%m-%d %H:%M"),  # Date
            full_descr,                          # Text
            "",                                  # Link
            str((Path("final") / f"{file_path.stem}_subtitled.mp4").resolve()),
            title,                               # Title
            platform,                            # Label(s)
            "", "", "",                          # Alt, Comment, Board/Album
            ("Reel" if platform.startswith("Instagram") else
            "Short" if platform.startswith("YouTube") else ""),
            "",                                  # CTA
            "FALSE",                             # Reminder
        ]
    )

wb.save(excel_path)
print(f"[OK] Geüpdatet schema in {excel_path}")
